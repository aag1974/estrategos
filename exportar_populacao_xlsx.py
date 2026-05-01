"""
Exporta a tabela do menu "População" do dashboard para Excel.
Fonte: outputs_fase2/tabela_mestre_ra.csv (PDAD 2021).
Saída: outputs_export/populacao_ra.xlsx
"""
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "outputs_fase2" / "tabela_mestre_ra.csv"
OUT_DIR = ROOT / "outputs_export"
OUT_DIR.mkdir(exist_ok=True)
OUT = OUT_DIR / "populacao_ra.xlsx"

COLS = [
    ("RA_NOME",                  "Região Administrativa", None,    "text"),
    ("DOM_renda_pc_media",       "Renda per capita (R$)", 0,       "money"),
    ("MOR_pct_superior",         "% Superior (16+)",      1,       "pct"),
    ("MOR_pct_sem_fund",         "% Sem fundamental",     1,       "pct"),
    ("DOM_pct_classe_AB",        "% Classe A/B",          1,       "pct"),
    ("DOM_pct_classe_DE",        "% Classe D/E",          1,       "pct"),
    ("DOM_pct_inseg_alimentar",  "% Inseg. alimentar",    1,       "pct"),
    ("MOR_pct_servidor_total",   "% Func. público",       1,       "pct"),
    ("MOR_pct_servidor_fed",     "% Serv. federal",       1,       "pct"),
    ("MOR_pct_privado",          "% Privado",             1,       "pct"),
    ("MOR_pct_conta_propria",    "% Autônomo",            1,       "pct"),
    ("MOR_pct_desocupado",       "% Desempregado",        1,       "pct"),
    ("MOR_pct_beneficio_social", "% Bolsa Família/BPC",   1,       "pct"),
    ("MOR_pct_plano_saude",      "% Plano de saúde",      1,       "pct"),
    ("MOR_pct_jovem_pop",        "% Jovens 16-24",        1,       "pct"),
    ("MOR_pct_idoso_pop",        "% Idosos 60+",          1,       "pct"),
    ("MOR_pct_nativo_df",        "% Nativo DF",           1,       "pct"),
    ("MOR_pct_migrante",         "% Migrante",            1,       "pct"),
]

df = pd.read_csv(SRC)
df = df[df["RA_NOME"].notna()].copy()

src_cols = [c for c, _, _, _ in COLS]
out = df[src_cols].copy()
out.columns = [label for _, label, _, _ in COLS]
out = out.sort_values("Região Administrativa").reset_index(drop=True)

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name="População PDAD 2021", index=False)
    ws = writer.sheets["População PDAD 2021"]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="854F0B")
    for col_idx, (_, label, _, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_, _, decimals, kind) in enumerate(COLS, start=1):
        letter = get_column_letter(col_idx)
        if kind == "money":
            fmt = '"R$" #,##0'
        elif kind == "pct":
            fmt = f'0.{"0"*decimals}"%"' if decimals else '0"%"'
        else:
            fmt = None
        if fmt:
            for row in range(2, len(out) + 2):
                ws.cell(row=row, column=col_idx).number_format = fmt
        if kind == "text":
            ws.column_dimensions[letter].width = 26
        else:
            ws.column_dimensions[letter].width = 14

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

print(f"OK: {OUT} ({len(out)} regiões)")
